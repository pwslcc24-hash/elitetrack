"""Load elite runner data from the scraped Excel workbook."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_XLSX = Path(__file__).parent / "elite_runners_tfrrs.xlsx"


def _row_dict(headers: list[str], values: tuple) -> dict[str, Any]:
    return {h: (v if v is not None else "") for h, v in zip(headers, values)}


def load_data(xlsx_path: Path | None = None) -> dict[str, Any]:
    path = xlsx_path or DEFAULT_XLSX
    if not path.is_file():
        raise FileNotFoundError(f"Data file not found: {path}. Run scrape_elite_runners.py first.")

    wb = load_workbook(path, read_only=True, data_only=True)

    marks_ws = wb["All Qualifying Marks"]
    marks_iter = marks_ws.iter_rows(values_only=True)
    mark_headers = list(next(marks_iter))
    marks = [_row_dict(mark_headers, row) for row in marks_iter if row and row[0]]

    ath_ws = wb["Unique Athletes"]
    ath_iter = ath_ws.iter_rows(values_only=True)
    ath_headers = list(next(ath_iter))
    athletes = [_row_dict(ath_headers, row) for row in ath_iter if row and row[0]]

    wb.close()

    gender_key = {"Men": "m", "Women": "f"}
    for a in athletes:
        a["_gender"] = gender_key.get(a["Gender"], "")
    for m in marks:
        m["_gender"] = gender_key.get(m["Gender"], "")

    return {
        "athletes": athletes,
        "marks": marks,
        "source": str(path.resolve()),
    }
