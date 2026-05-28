#!/usr/bin/env python3
"""
Scrape TFRRS division qualifying lists (2024–2026, outdoor + indoor) and export
athletes meeting elite time standards to Excel.
"""

from __future__ import annotations

import argparse
import re
import time
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE = "https://tf.tfrrs.org"
SESSION = requests.Session()
SESSION.headers.update(
    {
        "User-Agent": "tfrrs-elite-runners/1.0 (educational roster export)",
        "Accept": "text/html,application/xhtml+xml",
    }
)

# (season_label, division, list_id, slug)
LIST_SOURCES: List[Tuple[str, str, int, str]] = [
    # 2026 outdoor
    ("2026 outdoor", "NCAA D1", 5602, "2026_NCAA_Division_I_Outdoor_Qualifying_FINAL"),
    ("2026 outdoor", "NCAA D2", 5603, "2026_NCAA_Division_II_Outdoor_Qualifying_FINAL"),
    ("2026 outdoor", "NCAA D3", 5604, "2026_NCAA_Division_III_Outdoor_Qualifying_FINAL"),
    ("2026 outdoor", "NAIA", 5605, "2026_NAIA_Outdoor_Qualifying_List_FINAL"),
    ("2026 outdoor", "NJCAA", 5606, "NJCAA_All_Schools_Outdoor_Performance_List"),
    ("2026 outdoor", "NWAC", 5607, "2026_NWAC_Outdoor_Track__Field_Qualifying_List"),
    ("2026 outdoor", "NCCAA", 5608, "2026_NCCAA_Outdoor_Qualifying_List"),
    ("2026 outdoor", "USCAA", 5609, "2026_USCAA_Outdoor_Track__Field_List"),
    # 2025 outdoor
    ("2025 outdoor", "NCAA D1", 5018, "2025_NCAA_Division_I_Outdoor_Qualifying_List"),
    ("2025 outdoor", "NCAA D2", 5019, "2025_NCAA_Division_II_Outdoor_Qualifying_List"),
    ("2025 outdoor", "NCAA D3", 5020, "2025_NCAA_Division_III_Outdoor_Qualifying_List"),
    ("2025 outdoor", "NAIA", 5021, "2025_NAIA_Outdoor_Qualifying_List_FINAL"),
    ("2025 outdoor", "NJCAA", 5022, "NJCAA_All_Schools_Outdoor_Performance_List"),
    ("2025 outdoor", "NWAC", 5023, "2025_NWAC_Outdoor_Track__Field_Qualifying_List"),
    ("2025 outdoor", "NCCAA", 5024, "2025_NCCAA_Outdoor_Qualifying_List"),
    ("2025 outdoor", "USCAA", 5025, "2025_USCAA_Outdoor_Track__Field_List"),
    # 2024 outdoor
    ("2024 outdoor", "NCAA D1", 4479, "2024_NCAA_Division_I_Outdoor_Qualifying_List"),
    ("2024 outdoor", "NCAA D2", 4480, "2024_NCAA_Division_II_Outdoor_Qualifying_List"),
    ("2024 outdoor", "NCAA D3", 4481, "2024_NCAA_Division_III_Outdoor_Qualifying_List"),
    ("2024 outdoor", "NAIA", 4482, "2024_NAIA_Outdoor_Qualifying_List_FINAL"),
    ("2024 outdoor", "NJCAA", 4483, "NJCAA_All_Schools_Outdoor_Performance_List"),
    ("2024 outdoor", "NWAC", 4484, "2024_NWAC_Outdoor_Track__Field_Qualifying_List"),
    ("2024 outdoor", "NCCAA", 4485, "2024_NCCAA_Outdoor_Qualifying_List"),
    ("2024 outdoor", "USCAA", 4486, "2024_USCAA_Outdoor_Track__Field_List"),
    # 2026 indoor (2025–26 academic year)
    ("2026 indoor", "NAIA", 5351, "2025_2026_NAIA_Indoor_Qualifying_List_FINAL"),
    ("2026 indoor", "NCAA D1", 5352, "2025_2026_NCAA_Division_I_Indoor_List_FINAL"),
    ("2026 indoor", "NCAA D2", 5353, "2025_2026_NCAA_Division_II_Indoor_List_FINAL"),
    ("2026 indoor", "NCAA D3", 5354, "2025_2026_NCAA_Division_III_Indoor_List_FINAL"),
    ("2026 indoor", "NJCAA", 5355, "2025_2026_NJCAA_Indoor_Qualifying_List"),
    ("2026 indoor", "NCCAA", 5356, "2025_2026_NCCAA_Indoor_Qualifying"),
    ("2026 indoor", "NWAC", 5357, "2025_2026_NWAC_Indoor_Qualifying_List"),
    ("2026 indoor", "USCAA", 5358, "2025_2026_USCAA_Indoor_Qualifying_List"),
    # 2025 indoor (2024–25 academic year)
    ("2025 indoor", "NAIA", 4866, "2024_2025_NAIA_Indoor_Qualifying_List_FINAL"),
    ("2025 indoor", "NCAA D1", 4867, "2024_2025_NCAA_Division_I_Indoor_List_FINAL"),
    ("2025 indoor", "NCAA D2", 4868, "2024_2025_NCAA_Division_II_Indoor_List_FINAL"),
    ("2025 indoor", "NCAA D3", 4869, "2024_2025_NCAA_Division_III_Indoor_List_FINAL"),
    ("2025 indoor", "NJCAA", 4870, "2024_2025_NJCAA_Indoor_Qualifying_List"),
    ("2025 indoor", "NCCAA", 4871, "2024_2025_NCCAA_Indoor_Qualifying"),
    ("2025 indoor", "NWAC", 4872, "2024_2025_NWAC_Indoor_Qualifying_List"),
    ("2025 indoor", "USCAA", 4873, "2024_2025_USCAA_Indoor_Qualifying_List"),
    # 2024 indoor (2023–24 academic year)
    ("2024 indoor", "NAIA", 4363, "2023_2024_NAIA_Indoor_Qualifying_List_FINAL"),
    ("2024 indoor", "NCAA D1", 4364, "2023_2024_NCAA_Div_I_Indoor_Qualifying_FINAL"),
    ("2024 indoor", "NCAA D2", 4365, "2023_2024_NCAA_Division_II_Indoor_Qualifying_FINAL"),
    ("2024 indoor", "NCAA D3", 4366, "2023_2024_NCAA_Division_III_Indoor_Qualifying_FINAL"),
    ("2024 indoor", "NJCAA", 4367, "2023_2024_NJCAA_Indoor_Qualifying_List"),
    ("2024 indoor", "NCCAA", 4368, "2023_2024_NCCAA_Indoor_Qualifying"),
    ("2024 indoor", "NWAC", 4369, "2023_2024_NWAC_Indoor_Qualifying_List"),
    ("2024 indoor", "USCAA", 4370, "2023_2024_USCAA_Indoor_Qualifying_List"),
]

# User wrote "20k" — interpreted as 10,000m.
STANDARDS = {
    "f": {
        "10000": 34 * 60,
        "5000": 16 * 60 + 30,
        "steeple": 10 * 60 + 30,
        "3000": 9 * 60 + 30,
        "mile": 5 * 60,
        "1500": 4 * 60 + 30,
        "800": 2 * 60 + 10,
    },
    "m": {
        "10000": 29 * 60,
        "5000": 14 * 60,
        "steeple": 9 * 60,
        "3000": 8 * 60 + 5,
        "mile": 4 * 60 + 3,
        "1500": 3 * 60 + 41,
        "800": 1 * 60 + 50,
    },
}

EVENT_LABELS = {
    "10000": "10000m",
    "5000": "5000m",
    "steeple": "3000m Steeplechase",
    "3000": "3000m",
    "mile": "Mile",
    "1500": "1500m",
    "800": "800m",
}

OUTDOOR_EVENTS = {"10000", "5000", "steeple", "1500", "800"}
INDOOR_EVENTS = {"3000", "mile", "1500", "800", "5000", "steeple"}

TIME_RE = re.compile(r"^\d{1,2}:\d{2}(\.\d{2})?$")
GENDER_LABEL = {"m": "Men", "f": "Women"}


def parse_time_seconds(raw: str) -> Optional[float]:
    raw = raw.strip()
    if not TIME_RE.match(raw):
        return None
    parts = raw.split(":")
    if len(parts) != 2:
        return None
    minutes = int(parts[0])
    seconds = float(parts[1])
    return minutes * 60 + seconds


def normalize_event_name(heading: str) -> Optional[str]:
    h = heading.lower().replace(",", "")
    if "steeple" in h:
        return "steeple"
    if re.search(r"\b10\s*000\b|\b10000\b", h):
        return "10000"
    if re.search(r"\b5000\b", h):
        return "5000"
    if re.search(r"\b3000\b", h) and "steeple" not in h:
        return "3000"
    if "mile" in h and "relay" not in h:
        return "mile"
    if re.search(r"\b1500\b", h):
        return "1500"
    if re.search(r"\b800\b", h) and "relay" not in h:
        return "800"
    return None


def looks_like_time(text: str) -> bool:
    return parse_time_seconds(text) is not None


def allowed_events_for_season(season_label: str) -> Set[str]:
    if season_label.endswith("outdoor"):
        return OUTDOOR_EVENTS
    return INDOOR_EVENTS


@dataclass
class Performance:
    gender: str
    division: str
    season_label: str
    calendar_year: int
    event_key: str
    event_label: str
    time: str
    time_seconds: float
    athlete_name: str
    athlete_id: str
    profile_url: str
    school: str
    year: str
    meet: str
    meet_date: str
    list_id: int
    list_slug: str


@dataclass
class AthleteRecord:
    gender: str
    athlete_id: str
    athlete_name: str
    profile_url: str
    schools: Set[str] = field(default_factory=set)
    divisions: Set[str] = field(default_factory=set)
    season_labels: Set[str] = field(default_factory=set)
    marks: List[Performance] = field(default_factory=list)


def list_url(list_id: int, slug: str, gender: str) -> str:
    return f"{BASE}/lists/{list_id}/{slug}?gender={gender}&limit=500"


def fetch_html(url: str, delay: float) -> str:
    time.sleep(delay)
    resp = SESSION.get(url, timeout=90)
    resp.raise_for_status()
    return resp.text


def parse_list_page(
    html: str,
    *,
    gender: str,
    division: str,
    season_label: str,
    list_id: int,
    list_slug: str,
    allowed_events: Set[str],
) -> List[Performance]:
    soup = BeautifulSoup(html, "lxml")
    frame = soup.select_one("#list_data") or soup
    results: List[Performance] = []
    threshold_map = STANDARDS[gender]
    calendar_year = int(season_label.split()[0])

    for block in frame.select('.row[class*="standard_event"]'):
        heading_el = block.select_one("h2, h3, .panel-heading")
        if not heading_el:
            continue
        heading = heading_el.get_text(" ", strip=True)
        event_key = normalize_event_name(heading)
        if not event_key or event_key not in allowed_events:
            continue
        if event_key not in threshold_map:
            continue
        threshold = threshold_map[event_key]
        event_label = EVENT_LABELS[event_key]

        for row in block.select(".performance-list-row"):
            athlete_a = row.select_one(".col-athlete a[href*='/athletes/']")
            if not athlete_a:
                continue
            href = athlete_a.get("href", "")
            m = re.search(r"/athletes/(\d+)/", href)
            if not m:
                continue
            athlete_id = m.group(1)
            athlete_name = athlete_a.get_text(strip=True)
            profile_url = urljoin(BASE, href)

            team_el = row.select_one(".col-team")
            school = team_el.get_text(strip=True) if team_el else ""
            meet_el = row.select_one(".col-meet")
            meet = meet_el.get_text(strip=True) if meet_el else ""

            class_year = ""
            meet_date = ""
            mark = ""
            for col in row.select(".col-narrow"):
                text = col.get_text(strip=True)
                if not class_year and re.match(r"^(SR-4|JR-3|SO-2|FR-1)$", text):
                    class_year = text
                elif looks_like_time(text):
                    mark = text
                elif re.match(r"^[A-Z][a-z]{2}\s+\d{1,2},\s+\d{4}$", text):
                    meet_date = text

            if not mark:
                continue
            seconds = parse_time_seconds(mark)
            if seconds is None or seconds > threshold:
                continue

            results.append(
                Performance(
                    gender=gender,
                    division=division,
                    season_label=season_label,
                    calendar_year=calendar_year,
                    event_key=event_key,
                    event_label=event_label,
                    time=mark,
                    time_seconds=seconds,
                    athlete_name=athlete_name,
                    athlete_id=athlete_id,
                    profile_url=profile_url,
                    school=school,
                    year=class_year,
                    meet=meet,
                    meet_date=meet_date,
                    list_id=list_id,
                    list_slug=list_slug,
                )
            )

    return results


def scrape_all(delay: float) -> List[Performance]:
    all_perfs: List[Performance] = []
    total = len(LIST_SOURCES) * 2
    n = 0

    for season_label, division, list_id, slug in LIST_SOURCES:
        allowed = allowed_events_for_season(season_label)
        for gender in ("m", "f"):
            n += 1
            url = list_url(list_id, slug, gender)
            print(f"[{n}/{total}] {season_label} | {division} | {GENDER_LABEL[gender]}")
            html = fetch_html(url, delay)
            perfs = parse_list_page(
                html,
                gender=gender,
                division=division,
                season_label=season_label,
                list_id=list_id,
                list_slug=slug,
                allowed_events=allowed,
            )
            print(f"    -> {len(perfs)} qualifying marks")
            all_perfs.extend(perfs)

    return all_perfs


def aggregate(perfs: Iterable[Performance]) -> Dict[Tuple[str, str], AthleteRecord]:
    by_athlete: Dict[Tuple[str, str], AthleteRecord] = {}
    for p in perfs:
        key = (p.gender, p.athlete_id)
        if key not in by_athlete:
            by_athlete[key] = AthleteRecord(
                gender=p.gender,
                athlete_id=p.athlete_id,
                athlete_name=p.athlete_name,
                profile_url=p.profile_url,
            )
        rec = by_athlete[key]
        if p.school:
            rec.schools.add(p.school)
        rec.divisions.add(p.division)
        rec.season_labels.add(p.season_label)
        rec.marks.append(p)
    return by_athlete


def format_marks_summary(marks: List[Performance]) -> str:
    grouped: Dict[str, List[str]] = defaultdict(list)
    for m in sorted(marks, key=lambda x: (x.calendar_year, x.event_key, x.time_seconds)):
        grouped[m.event_label].append(
            f"{m.time} ({m.season_label}, {m.division}, {m.school or 'n/a'}, {m.meet_date or 'n/a'})"
        )
    parts = []
    for event in EVENT_LABELS.values():
        if event in grouped:
            parts.append(f"{event}: " + "; ".join(grouped[event]))
    return " | ".join(parts)


def autosize_columns(ws, max_row_sample: int = 300) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 12
        for row in range(1, min(ws.max_row, max_row_sample) + 1):
            val = ws[f"{letter}{row}"].value
            if val:
                max_len = max(max_len, min(len(str(val)), 80))
        ws.column_dimensions[letter].width = max_len + 2


def export_xlsx(
    perfs: List[Performance],
    athletes: Dict[Tuple[str, str], AthleteRecord],
    output: Path,
) -> None:
    wb = Workbook()

    ws_marks = wb.active
    ws_marks.title = "All Qualifying Marks"
    mark_headers = [
        "Gender",
        "Athlete",
        "TFRRS Profile",
        "Athlete ID",
        "School",
        "Division",
        "Season",
        "Event",
        "Time",
        "Class Year",
        "Meet",
        "Meet Date",
        "List ID",
    ]
    ws_marks.append(mark_headers)
    for cell in ws_marks[1]:
        cell.font = Font(bold=True)

    for p in sorted(
        perfs,
        key=lambda x: (x.gender, x.athlete_name, x.calendar_year, x.event_key, x.time_seconds),
    ):
        ws_marks.append(
            [
                GENDER_LABEL[p.gender],
                p.athlete_name,
                p.profile_url,
                p.athlete_id,
                p.school,
                p.division,
                p.season_label,
                p.event_label,
                p.time,
                p.year,
                p.meet,
                p.meet_date,
                p.list_id,
            ]
        )

    ws_ath = wb.create_sheet("Unique Athletes")
    ath_headers = [
        "Gender",
        "Athlete",
        "TFRRS Profile",
        "Athlete ID",
        "School(s)",
        "Division(s)",
        "Season(s)",
        "# Qualifying Marks",
        "Events Met",
        "All Qualifying Times",
    ]
    ws_ath.append(ath_headers)
    for cell in ws_ath[1]:
        cell.font = Font(bold=True)

    for (_gender, _aid), rec in sorted(
        athletes.items(),
        key=lambda item: (item[1].gender, item[1].athlete_name),
    ):
        events_met = sorted({m.event_label for m in rec.marks})
        ws_ath.append(
            [
                GENDER_LABEL[rec.gender],
                rec.athlete_name,
                rec.profile_url,
                rec.athlete_id,
                "; ".join(sorted(rec.schools)),
                "; ".join(sorted(rec.divisions)),
                "; ".join(sorted(rec.season_labels)),
                len(rec.marks),
                ", ".join(events_met),
                format_marks_summary(rec.marks),
            ]
        )

    ws_std = wb.create_sheet("Standards")
    ws_std.append(["Gender", "Event", "Standard (faster than)"])
    for cell in ws_std[1]:
        cell.font = Font(bold=True)
    for gender_name, key in [("Women", "f"), ("Men", "m")]:
        for event_key, seconds in STANDARDS[key].items():
            mins = int(seconds // 60)
            secs = seconds % 60
            ws_std.append([gender_name, EVENT_LABELS[event_key], f"{mins}:{secs:05.2f}"])

    ws_lists = wb.create_sheet("Lists Scraped")
    ws_lists.append(["Season", "Division", "List ID", "Slug", "URL"])
    for cell in ws_lists[1]:
        cell.font = Font(bold=True)
    for season_label, division, list_id, slug in LIST_SOURCES:
        ws_lists.append(
            [season_label, division, list_id, slug, f"{BASE}/lists/{list_id}/{slug}"]
        )

    for ws in (ws_marks, ws_ath, ws_std, ws_lists):
        autosize_columns(ws)

    wb.save(output)


def print_summary(perfs: List[Performance], athletes: Dict[Tuple[str, str], AthleteRecord], output: Path) -> None:
    by_season = Counter(p.season_label for p in perfs)
    by_year = Counter(p.calendar_year for p in perfs)

    print()
    print(f"Qualifying performances: {len(perfs)}")
    print(f"Unique athletes: {len(athletes)}")
    print(f"Wrote {output.resolve()}")
    print("\nMarks by season label:")
    for season in sorted(by_season.keys()):
        print(f"  {season}: {by_season[season]}")
    print("\nMarks by calendar year:")
    for year in sorted(by_year.keys()):
        print(f"  {year}: {by_year[year]}")
    print(f"\nList sources scraped: {len(LIST_SOURCES)} (×2 genders = {len(LIST_SOURCES)*2} requests)")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export TFRRS elite runners to Excel")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("elite_runners_tfrrs.xlsx"),
        help="Output .xlsx path",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.75,
        help="Seconds between TFRRS requests (be polite)",
    )
    args = parser.parse_args()

    perfs = scrape_all(args.delay)
    athletes = aggregate(perfs)
    export_xlsx(perfs, athletes, args.output)
    print_summary(perfs, athletes, args.output)


if __name__ == "__main__":
    main()
